"""Tests for :mod:`concinno.tools.builtin.read_attachment`.

Covers:

* Tool contract (``name`` / ``description`` / ``is_concurrency_safe``).
* xlsx dispatch → TSV with ``[Sheet: name]`` header.
* csv dispatch → TSV rows.
* Plain-text extension dispatch (json / jsonl / pdb / txt).
* Unknown-extension fallback: text if printable, structured error if binary.
* Missing file / directory / empty path / blocked device → ``error: ...``.
* Lazy openpyxl import failure path (monkeypatched) for the
  zero-runtime-deps contract — a clean ``error: ...`` string.

Output-shape assertions are deliberately loose (contains-based) so downstream
format tweaks do not break every test at once. Hard invariants — sheet
header presence, TSV cell separator, ``error:`` prefix on failure — are
strict.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from concinno.tools.builtin.read_attachment import (
    DEFAULT_HEAD_PREVIEW_CHARS,
    DEFAULT_MAX_INLINE_CHARS,
    DEFAULT_MAX_ROWS_PER_SHEET,
    ReadAttachmentError,
    ReadAttachmentTool,
    _resolve_max_inline_chars,
)

# ── Contract ─────────────────────────────────────────


def test_contract_attrs() -> None:
    t = ReadAttachmentTool()
    assert t.name == "read_attachment"
    assert t.is_concurrency_safe is True
    desc = t.description.lower()
    # Description must name what formats it handles + the anti-retry
    # guidance so weak models do not loop on binary formats.
    assert "xlsx" in desc
    assert "csv" in desc
    assert "plain" in desc or "text" in desc


# ── xlsx happy path ──────────────────────────────────


@pytest.fixture
def xlsx_path(tmp_path: Path) -> Path:
    """A two-sheet workbook: first sheet has data, second is empty."""
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Inventory"
    ws1.append(["Title", "Format", "Year"])
    ws1.append(["Time-Parking 2: Parallel Universe", "Blu-Ray", 1991])
    ws1.append(["Other Film", "DVD", 2005])
    ws2 = wb.create_sheet("Notes")  # left empty on purpose
    del ws2
    p = tmp_path / "inv.xlsx"
    wb.save(p)
    return p


def test_xlsx_rendering(xlsx_path: Path) -> None:
    t = ReadAttachmentTool()
    out = t.call(path=str(xlsx_path))
    # Sheet markers present.
    assert "[Sheet: Inventory]" in out
    assert "[Sheet: Notes]" in out
    # Data rows preserved, TSV-separated.
    assert "Title\tFormat\tYear" in out
    assert "Time-Parking 2: Parallel Universe\tBlu-Ray\t1991" in out
    # Empty sheet annotation.
    assert "[empty sheet]" in out


def test_xlsx_row_cap(tmp_path: Path) -> None:
    """Row cap fires and annotates the truncation."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    # Overshoot the cap by a healthy margin.
    for i in range(DEFAULT_MAX_ROWS_PER_SHEET + 20):
        ws.append([i, f"row-{i}"])
    p = tmp_path / "big.xlsx"
    wb.save(p)
    out = ReadAttachmentTool().call(path=str(p))
    assert f"Row cap {DEFAULT_MAX_ROWS_PER_SHEET}" in out


def test_xlsx_cell_sanitization(tmp_path: Path) -> None:
    """Tabs/newlines in a cell must be replaced to keep TSV framing."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["a\tb", "c\nd"])
    p = tmp_path / "mess.xlsx"
    wb.save(p)
    out = ReadAttachmentTool().call(path=str(p))
    # Row has exactly one tab separator between the two cells → no
    # extra tabs leaked from the cell content.
    data_line = [ln for ln in out.splitlines() if "a" in ln and "c" in ln][0]
    assert data_line.count("\t") == 1
    assert "\n" not in data_line


# ── csv ──────────────────────────────────────────────


def test_csv_rendering(tmp_path: Path) -> None:
    p = tmp_path / "data.csv"
    p.write_text("h1,h2,h3\nv1,v2,v3\n", encoding="utf-8")
    out = ReadAttachmentTool().call(path=str(p))
    assert "h1\th2\th3" in out
    assert "v1\tv2\tv3" in out


def test_csv_empty(tmp_path: Path) -> None:
    p = tmp_path / "empty.csv"
    p.write_text("", encoding="utf-8")
    out = ReadAttachmentTool().call(path=str(p))
    assert "[empty csv]" in out


# ── plain-text extensions ────────────────────────────


@pytest.mark.parametrize(
    "ext,payload",
    [
        (".txt", "hello world"),
        (".md", "# heading\n\nparagraph"),
        (".json", '{"a":1,"b":2}'),
        (".jsonld", '{"@context":"schema.org"}'),
        (".pdb", "ATOM      1  N   ALA A   1"),
        (".yaml", "a: 1\nb: 2"),
    ],
)
def test_plain_text_extensions(tmp_path: Path, ext: str, payload: str) -> None:
    p = tmp_path / f"file{ext}"
    p.write_text(payload, encoding="utf-8")
    out = ReadAttachmentTool().call(path=str(p))
    assert payload.splitlines()[0] in out


# ── Unknown extension / binary sniff ─────────────────


def test_unknown_extension_text_passes(tmp_path: Path) -> None:
    p = tmp_path / "mystery.xyz"
    p.write_text("plain ascii content", encoding="utf-8")
    out = ReadAttachmentTool().call(path=str(p))
    assert "plain ascii content" in out


def test_unknown_extension_binary_errors(tmp_path: Path) -> None:
    """NUL bytes in head → structured error, not garbage."""
    p = tmp_path / "mystery.xyz"
    p.write_bytes(b"\x00\x01\x02\xff" * 200)
    out = ReadAttachmentTool().call(path=str(p))
    assert out.startswith("error:")
    # Message must tell the model not to retry blindly.
    assert "binary" in out.lower() or "cannot" in out.lower()


# ── Failure modes ────────────────────────────────────


def test_empty_path() -> None:
    assert ReadAttachmentTool().call(path="").startswith("error:")


def test_missing_file(tmp_path: Path) -> None:
    out = ReadAttachmentTool().call(path=str(tmp_path / "nope.xlsx"))
    assert "not found" in out


def test_directory_refused(tmp_path: Path) -> None:
    out = ReadAttachmentTool().call(path=str(tmp_path))
    assert "not a regular file" in out


def test_unc_path_refused() -> None:
    assert "refusing" in ReadAttachmentTool().call(path="\\\\host\\share")


# ── Lazy-import failure contract ─────────────────────


def test_openpyxl_missing_returns_clean_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Simulate a runtime without openpyxl — tool must not raise."""
    import sys
    # Create a plausible .xlsx path. We do not need the bytes to be a
    # real workbook because the tool should fail at import before
    # touching the file.
    p = tmp_path / "dummy.xlsx"
    p.write_bytes(b"PK\x03\x04fake")

    # Hide openpyxl from the lazy import inside _read_xlsx.
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    out = ReadAttachmentTool().call(path=str(p))
    assert out.startswith("error:")
    assert "openpyxl" in out


def test_readattachmenterror_is_exception() -> None:
    # Type contract — raised internally, string-wrapped by ``call``.
    assert issubclass(ReadAttachmentError, Exception)


# ── Inline cap / summary fallback ────────────────────────


def test_small_file_under_cap_returns_full_content(tmp_path: Path) -> None:
    """Below the inline cap the tool behaves exactly as before."""
    p = tmp_path / "small.txt"
    body = "hello world\n" * 50  # ~600 chars — well under 15k default
    p.write_text(body)
    out = ReadAttachmentTool().call(path=str(p))
    assert out.startswith("hello world")
    assert "read_attachment: file too large" not in out


def test_large_file_over_cap_returns_summary(tmp_path: Path) -> None:
    """Over the inline cap the tool returns a short summary only."""
    p = tmp_path / "big.pdb"
    # Mimic a PDB: many short lines so the size-on-disk and the
    # rendered-chars both clear the cap.
    body = "ATOM  12345  N   ALA A   1      11.104  13.207  10.266\n" * 1000
    p.write_text(body)
    out = ReadAttachmentTool().call(
        path=str(p), max_inline_chars=2000,
    )
    # Summary markers present.
    assert "read_attachment: file too large" in out
    assert "path:" in out
    assert str(p) in out
    assert "size_bytes:" in out
    assert "extension: .pdb" in out
    assert "how_to_process" in out
    assert "python_exec" in out
    # Head preview present but bounded.
    assert "head_preview" in out
    assert "ATOM" in out
    # Summary itself must not exceed a reasonable budget.
    assert len(out) < 2500


def test_large_file_summary_steers_to_python_exec(tmp_path: Path) -> None:
    """Summary actively routes the model to ``python_exec`` with
    ``open(path)`` — the whole point of the inline cap."""
    p = tmp_path / "big.log"
    p.write_text("line\n" * 5000)
    out = ReadAttachmentTool().call(
        path=str(p), max_inline_chars=500,
    )
    assert "python_exec" in out
    # Explicit anti-retry guidance — without this weak models loop on
    # re-reading the same path.
    assert "Do NOT retry" in out or "retry" in out.lower()


def test_max_inline_chars_kwarg_overrides_default(tmp_path: Path) -> None:
    p = tmp_path / "medium.txt"
    body = "a" * 5000
    p.write_text(body)
    # With a very small cap we get a summary.
    assert "file too large" in ReadAttachmentTool().call(
        path=str(p), max_inline_chars=1000,
    )
    # With a huge cap we get the full content.
    out = ReadAttachmentTool().call(
        path=str(p), max_inline_chars=10_000_000,
    )
    assert "file too large" not in out
    assert out.startswith("a" * 100)


def test_max_inline_chars_zero_disables_cap(tmp_path: Path) -> None:
    """Setting the cap to 0 returns the full rendering. Useful for
    debug deploys or test runs where truncation would hide signal."""
    p = tmp_path / "text.txt"
    body = "x" * 50_000
    p.write_text(body)
    out = ReadAttachmentTool().call(path=str(p), max_inline_chars=0)
    assert "file too large" not in out
    assert out.startswith("x" * 100)


def test_env_override_applies_when_kwarg_none(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("CONCINNO_READ_ATTACHMENT_MAX_INLINE_CHARS", "500")
    p = tmp_path / "log.txt"
    p.write_text("line\n" * 1000)
    out = ReadAttachmentTool().call(path=str(p))
    assert "file too large" in out


def test_env_override_malformed_falls_back_to_default(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Garbage env value must not silently disable the guard."""
    monkeypatch.setenv("CONCINNO_READ_ATTACHMENT_MAX_INLINE_CHARS", "not-an-int")
    assert _resolve_max_inline_chars() == DEFAULT_MAX_INLINE_CHARS


def test_env_override_negative_falls_back_to_default(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("CONCINNO_READ_ATTACHMENT_MAX_INLINE_CHARS", "-42")
    assert _resolve_max_inline_chars() == DEFAULT_MAX_INLINE_CHARS


def test_env_override_zero_is_honoured(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("CONCINNO_READ_ATTACHMENT_MAX_INLINE_CHARS", "0")
    assert _resolve_max_inline_chars() == 0


def test_env_override_positive_integer(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("CONCINNO_READ_ATTACHMENT_MAX_INLINE_CHARS", "7500")
    assert _resolve_max_inline_chars() == 7500


def test_head_preview_bounded(tmp_path: Path) -> None:
    """The head preview obeys :data:`DEFAULT_HEAD_PREVIEW_CHARS` even when
    the underlying rendering is enormous — cheap guardrail against a
    regression where the summary silently becomes almost as large as
    the original."""
    p = tmp_path / "big.pdb"
    p.write_text("A" * 100_000)
    out = ReadAttachmentTool().call(path=str(p), max_inline_chars=2000)
    # Find the head_preview block and verify it stays under the cap.
    idx = out.find("head_preview")
    assert idx >= 0
    preview_block = out[idx:]
    # Preview content is everything after the "chars):\n" marker up to
    # the terminator.
    start = preview_block.find("chars):\n")
    end = preview_block.find("[end of read_attachment summary]")
    assert start > 0
    assert end > start
    preview_text = preview_block[start + len("chars):\n"):end].rstrip()
    assert len(preview_text) <= DEFAULT_HEAD_PREVIEW_CHARS + 5


def test_binary_rejection_unaffected_by_cap(tmp_path: Path) -> None:
    """Binary sniff short-circuits before the cap — error string is
    returned, not a summary. Keeps the failure taxonomy clean."""
    p = tmp_path / "blob.bin"
    # A head that looks binary (NUL byte is a strong binary signal
    # inside _looks_binary).
    p.write_bytes(b"\x00\x01\x02\x03" * 1000 + b"A" * 50000)
    out = ReadAttachmentTool().call(
        path=str(p), max_inline_chars=500,
    )
    assert out.startswith("error:")
    assert "file too large" not in out
